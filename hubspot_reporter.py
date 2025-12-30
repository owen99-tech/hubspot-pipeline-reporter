#!/usr/bin/env python3
"""
HubSpot Pipeline Reporter
Automatically fetches pipeline data from HubSpot and exports to Excel
"""

import os
import sys
from datetime import datetime
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

class HubSpotReporter:
    def __init__(self, api_key=None):
        self.api_key = api_key or os.getenv('HUBSPOT_API_KEY')
        if not self.api_key:
            raise ValueError("HubSpot API key not found. Please set HUBSPOT_API_KEY in .env file")
        
        self.base_url = "https://api.hubapi.com"
        self.headers = {
            'Authorization': f'Bearer {self.api_key}',
            'Content-Type': 'application/json'
        }
    
    def get_pipelines(self):
        """
        Fetch all available pipelines
        """
        try:
            url = f"{self.base_url}/crm/v3/pipelines/deals"
            response = requests.get(url, headers=self.headers)
            response.raise_for_status()
            return response.json()['results']
        except requests.exceptions.RequestException as e:
            print(f"Error fetching pipelines: {e}")
            sys.exit(1)
    
    def get_pipeline_deals(self, pipeline_id):
        """
        Fetch all deals for a specific pipeline
        """
        try:
            url = f"{self.base_url}/crm/v3/objects/deals"
            params = {
                'properties': 'dealname,amount,dealstage,closedate,pipeline,createdate',
                'limit': 100
            }
            
            deals = []
            after = None
            
            while True:
                if after:
                    params['after'] = after
                
                response = requests.get(url, headers=self.headers, params=params)
                response.raise_for_status()
                data = response.json()
                
                # Filter deals by pipeline
                for deal in data.get('results', []):
                    if deal['properties'].get('pipeline') == pipeline_id:
                        deals.append(deal)
                
                # Check for pagination
                paging = data.get('paging', {})
                if 'next' in paging:
                    after = paging['next'].get('after')
                else:
                    break
            
            return deals
        except requests.exceptions.RequestException as e:
            print(f"Error fetching deals: {e}")
            return []
    
    def export_to_excel(self, pipeline_name, deals, output_dir='reports'):
        """
        Export pipeline deals to Excel file
        """
        # Create output directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = pipeline_name[:31]  # Excel sheet name limit
        
        # Define headers
        headers = ['Deal Name', 'Amount', 'Stage', 'Close Date', 'Created Date']
        
        # Style headers
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        # Write headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        # Write deal data
        for row, deal in enumerate(deals, start=2):
            props = deal['properties']
            ws.cell(row=row, column=1).value = props.get('dealname', 'N/A')
            ws.cell(row=row, column=2).value = props.get('amount', 'N/A')
            ws.cell(row=row, column=3).value = props.get('dealstage', 'N/A')
            ws.cell(row=row, column=4).value = props.get('closedate', 'N/A')
            ws.cell(row=row, column=5).value = props.get('createdate', 'N/A')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{output_dir}/{pipeline_name.replace(' ', '_')}_{timestamp}.xlsx"
        
        # Save workbook
        wb.save(filename)
        print(f"Report saved: {filename}")
        return filename

def main():
    """
    Main execution function
    """
    try:
        # Initialize reporter
        reporter = HubSpotReporter()
        
        # Get all pipelines
        pipelines = reporter.get_pipelines()
        print(f"Found {len(pipelines)} pipeline(s)")
        
        # For this example, we'll export the first pipeline
        # You can modify this to select a specific pipeline
        if pipelines:
            pipeline = pipelines[0]
            pipeline_id = pipeline['id']
            pipeline_name = pipeline['label']
            
            print(f"\nFetching deals for pipeline: {pipeline_name}")
            deals = reporter.get_pipeline_deals(pipeline_id)
            print(f"Found {len(deals)} deal(s)")
            
            # Export to Excel
            if deals:
                reporter.export_to_excel(pipeline_name, deals)
            else:
                print("No deals found in this pipeline")
        else:
            print("No pipelines found")
    
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
