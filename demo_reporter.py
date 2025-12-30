#!/usr/bin/env python3
"""
Demo HubSpot Pipeline Reporter
Generates a sample Excel report with dummy pipeline data
No API connection required - perfect for testing the output format
"""

import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import random

def generate_dummy_deals():
    """
    Generate realistic dummy deal data
    """
    companies = [
        "Acme Corp", "TechStart Inc", "Global Solutions", "Innovation Labs",
        "Enterprise Co", "StartupXYZ", "Digital Ventures", "CloudTech",
        "DataSystems Ltd", "FutureTech", "SmartBiz", "NextGen Solutions",
        "Quantum Corp", "Velocity Inc", "Synergy Partners"
    ]
    
    stages = [
        "Qualified Lead", "Meeting Scheduled", "Proposal Sent", 
        "Negotiation", "Closed Won", "Closed Lost"
    ]
    
    deals = []
    base_date = datetime.now()
    
    for i in range(15):
        # Generate random dates
        created_days_ago = random.randint(10, 90)
        close_days_future = random.randint(-10, 60)
        
        created_date = (base_date - timedelta(days=created_days_ago)).strftime('%Y-%m-%d')
        close_date = (base_date + timedelta(days=close_days_future)).strftime('%Y-%m-%d')
        
        deal = {
            'properties': {
                'dealname': f"{random.choice(companies)} - {random.choice(['Q1', 'Q2', 'Q3', 'Q4'])} Deal",
                'amount': f"${random.randint(5000, 150000):,}",
                'dealstage': random.choice(stages),
                'closedate': close_date,
                'createdate': created_date
            }
        }
        deals.append(deal)
    
    return deals

def export_to_excel(pipeline_name, deals, output_dir='demo_reports'):
    """
    Export pipeline deals to Excel file (same format as main script)
    """
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = pipeline_name[:31]  # Excel sheet name limit
    
    # Define headers
    headers = ['Deal Name', 'Amount', 'Stage', 'Close Date', 'Created Date']
    
    # Style headers
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    
    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    
    # Write deal data
    for row, deal in enumerate(deals, start=2):
        props = deal['properties']
        ws.cell(row=row, column=1).value = props.get('dealname', 'N/A')
        ws.cell(row=row, column=2).value = props.get('amount', 'N/A')
        ws.cell(row=row, column=3).value = props.get('dealstage', 'N/A')
        ws.cell(row=row, column=4).value = props.get('closedate', 'N/A')
        ws.cell(row=row, column=5).value = props.get('createdate', 'N/A')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{output_dir}/{pipeline_name.replace(' ', '_')}_{timestamp}.xlsx"
    
    # Save workbook
    wb.save(filename)
    print(f"\n{'='*60}")
    print(f"Demo Report Generated Successfully!")
    print(f"{'='*60}")
    print(f"\nFile Location: {filename}")
    print(f"Total Deals: {len(deals)}")
    print(f"\nThis is a sample of what your HubSpot reports will look like.")
    print(f"Open the file to see the formatted Excel output.\n")
    
    return filename

def main():
    """
    Main execution - generate demo report
    """
    print("\n" + "="*60)
    print("HubSpot Pipeline Reporter - DEMO MODE")
    print("="*60)
    print("\nGenerating sample report with dummy data...\n")
    
    # Generate dummy data
    pipeline_name = "Sales_Pipeline_Demo"
    dummy_deals = generate_dummy_deals()
    
    # Export to Excel
    export_to_excel(pipeline_name, dummy_deals)
    
    print("\nNote: This demo uses randomly generated data.")
    print("The actual script will fetch real data from your HubSpot account.\n")

if __name__ == "__main__":
    main()
